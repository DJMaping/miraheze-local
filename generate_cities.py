"""Generate a '== Largest cities ==' section for each canonical (1765) country page.

Data sources:
  - data/Largest Cities by Population.xlsx  (sheet 'cities'): Country, Name, population
  - data/*Cities*.xlsx (per-region, one sheet per country): full roster, no population

Merge per country, populations where known else N/A, sort smallest->largest,
N/A cities alphabetical at the bottom.
"""
import openpyxl, glob, os, re, sys, collections

DATA = r"C:\Users\danny\Documents\miraheze-local\data"
PAGES = r"C:\Users\danny\Documents\miraheze-local\pages\Main"
START = "<!-- LARGEST-CITIES-AUTO START -->"
END = "<!-- LARGEST-CITIES-AUTO END -->"

# Spreadsheet sheet name -> actual wiki page title (spelling mismatches in the data)
ALIAS = {
    "Praesy": "Praesyu",
    "Yihnurga": "Yihnurda",
    "Eteretes": "Etretes",
    "Isubul": "Isuibul",
    "Oryreain": "Oyreain",
    "Syliaduun": "Slyiaduun",
}

def norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

# Some roster cells wrap the name in literal quotes (straight or curly). Strip them.
_QUOTES = '"“”‘’\''
def clean_name(s):
    return str(s).strip().strip(_QUOTES).strip()

def load_data():
    # populations: country -> {norm_city: (display_name, pop_int)}
    pop = collections.defaultdict(dict)
    wb = openpyxl.load_workbook(os.path.join(DATA, "Largest Cities by Population.xlsx"),
                               read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[1] is None or r[2] is None:
            continue
        country, city, p = str(r[1]).strip(), clean_name(r[2]), r[3]
        if isinstance(p, (int, float)) and p > 0:
            pop[country][norm(city)] = (city, int(p))
    wb.close()

    # roster (no population): country -> {norm_city: display_name}
    roster = collections.defaultdict(dict)
    for f in glob.glob(os.path.join(DATA, "*Cities*.xlsx")):
        if "Largest" in f:
            continue
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sh in wb.sheetnames:
            country = sh.strip()
            for r in wb[sh].iter_rows(min_row=2, values_only=True):
                if r and r[0] is not None and clean_name(r[0]):
                    city = clean_name(r[0])
                    roster[country].setdefault(norm(city), city)
        wb.close()
    return pop, roster

def merge_country(country, pop, roster):
    """Return sorted list of (display_name, pop_or_None)."""
    merged = {}  # norm -> (name, pop)
    for nc, name in roster.get(country, {}).items():
        merged[nc] = (name, None)
    for nc, (name, p) in pop.get(country, {}).items():
        merged[nc] = (name, p)  # population source wins (name + number)
    # Largest first; cities without a population ranked after, alphabetical.
    with_pop = sorted([v for v in merged.values() if v[1] is not None], key=lambda x: (-x[1], x[0]))
    no_pop = sorted([v for v in merged.values() if v[1] is None], key=lambda x: x[0].lower())
    return with_pop + no_pop, len(with_pop), len(no_pop)

def slug(country):
    return re.sub(r'[^a-z0-9]+', '-', country.lower()).strip('-')

TOP_N = 20  # cap the ranked box at the 20 largest, like the Wikipedia template

def build_section(country, cities):
    # cities: largest-first (pop desc), then N/A alphabetical.
    pop_cities = [(n, p) for (n, p) in cities if p is not None][:TOP_N]
    no_pop = [(n, p) for (n, p) in cities if p is None]
    lines = [START, "== Largest cities =="]

    # No population data anywhere -> simple alphabetical list of city names.
    if not pop_cities:
        lines.append(f"Cities in [[{country}]] (population figures not yet available):")
        for name, _ in no_pop:
            lines.append(f"* [[{name}]]")
        lines.append(END)
        return "\n".join(lines)

    caption = (
        f'|+ Largest cities in [[{country}]]<br /><small>Population: '
        f'World Population Prospects 1765<ref name="wpp1765">{{{{cite web'
        f'|url=https://worlddataunion.org/prospects/1765/{slug(country)}'
        f'|title=World Population Prospects 1765: {country}|publisher=World Data Union}}}}</ref></small>')
    lines.append('{| class="wikitable" style="text-align:center"')
    lines.append(caption)
    n = len(pop_cities)
    if n <= 10:
        lines.append("! Rank !! City !! Population")
        for i, (name, p) in enumerate(pop_cities, 1):
            lines.append("|-")
            lines.append(f"| {i} || [[{name}]] || {p:,}")
    else:
        # Two columns like the image: ranks 1-10 left, 11-N right.
        lines.append("! Rank !! City !! Population !! Rank !! City !! Population")
        for i in range(10):
            lname, lp = pop_cities[i]
            left = f"{i + 1} || [[{lname}]] || {lp:,}"
            ri = 10 + i
            if ri < n:
                rname, rp = pop_cities[ri]
                right = f"{ri + 1} || [[{rname}]] || {rp:,}"
            else:
                right = " || || "
            lines.append("|-")
            lines.append(f"| {left} || {right}")
    lines.append("|}")
    lines.append(END)
    return "\n".join(lines)

def insert_section(text, section):
    # Idempotent: replace existing auto block if present.
    if START in text and END in text:
        pre = text[:text.index(START)]
        post = text[text.index(END) + len(END):]
        return pre + section + post
    # Otherwise insert before '== References ==' (case-insensitive), else before first
    # category link, else append.
    m = re.search(r'(?im)^==\s*References\s*==', text)
    if m:
        idx = m.start()
        return text[:idx] + section + "\n\n" + text[idx:]
    m = re.search(r'(?im)^\[\[Category:', text)
    if m:
        idx = m.start()
        return text[:idx] + section + "\n\n" + text[idx:]
    return text.rstrip() + "\n\n" + section + "\n"

def page_path(country):
    return os.path.join(PAGES, country.replace(" ", "_") + ".wiki")

def target_countries():
    countries = set()
    for f in glob.glob(os.path.join(DATA, "*Cities*.xlsx")):
        if "Largest" in f:
            continue
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sh in wb.sheetnames:
            countries.add(sh.strip())
        wb.close()
    return sorted(countries)

def main():
    pop, roster = load_data()
    mode = sys.argv[1] if len(sys.argv) > 1 else "dry"
    if mode == "dry":
        for country in (sys.argv[2:] or ["Acetoa"]):
            cities, nwp, nnp = merge_country(country, pop, roster)
            print(f"\n##### {country}: {len(cities)} cities ({nwp} with pop, {nnp} N/A) #####")
            print(build_section(country, cities))
        return
    if mode == "apply":
        applied, skipped = [], []
        for country in target_countries():
            disp = ALIAS.get(country, country)  # page title may differ from sheet name
            path = page_path(disp)
            if not os.path.exists(path):
                skipped.append(country)
                continue
            cities, _, _ = merge_country(country, pop, roster)
            if not cities:
                skipped.append(country + " (no cities)")
                continue
            with open(path, "r", encoding="utf-8") as fh:
                text = fh.read()
            new = insert_section(text, build_section(disp, cities))
            with open(path, "w", encoding="utf-8", newline="\n") as fh:
                fh.write(new)
            applied.append(country)
        print(f"APPLIED to {len(applied)} pages.")
        print(f"SKIPPED ({len(skipped)}): {skipped}")

if __name__ == "__main__":
    main()
