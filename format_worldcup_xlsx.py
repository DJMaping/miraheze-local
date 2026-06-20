#!/usr/bin/env python3
"""Render data/FLLA_World_Cup_Players.csv as a styled .xlsx.

Produces a presentation-ready workbook (real table with banded rows, a bold
colour-block header, frozen header row, an auto-filter, sized columns, centred
numeric columns and a green colour-scale on Total goals) so it looks like a
proper table rather than a flat CSV grid. The styling is applied as plain cell
formatting too, so it survives conversion to a native Google Sheet on upload.

Usage:  python format_worldcup_xlsx.py
"""

import csv
import os

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
CSV_IN = os.path.join(HERE, "data", "FLLA_World_Cup_Players.csv")
XLSX_OUT = os.path.join(HERE, "data", "FLLA_World_Cup_Players.xlsx")

# Palette (FLLA pitch green)
HEADER_FILL = PatternFill("solid", fgColor="0B6B3A")
BAND_FILL = PatternFill("solid", fgColor="EAF4EE")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(color="20303A", size=10)
THIN = Side(style="thin", color="D6E3DA")

CENTRE_COLS = {"Position", "World Cups", "Total goals", "Goals by World Cup"}
MAX_WIDTH = 46


def main():
    with open(CSV_IN, encoding="utf-8", newline="") as fh:
        rows = list(csv.reader(fh))
    headers, data = rows[0], rows[1:]
    ncol = len(headers)
    nrow = len(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "World Cup Players"
    ws.sheet_view.showGridLines = False

    # Header
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="084D29"))

    centre_idx = {i for i, h in enumerate(headers) if h in CENTRE_COLS}

    # Body
    for r, record in enumerate(data, start=2):
        banded = (r % 2 == 0)
        for c in range(1, ncol + 1):
            raw = record[c - 1] if c - 1 < len(record) else ""
            value = int(raw) if (headers[c - 1] == "Total goals" and raw.isdigit()) else raw
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = CELL_FONT
            cell.border = Border(bottom=THIN)
            if banded:
                cell.fill = BAND_FILL
            cell.alignment = Alignment(
                horizontal="center" if (c - 1) in centre_idx else "left",
                vertical="center",
                wrap_text=(headers[c - 1] in ("Goals by World Cup", "Awards", "Notes")),
            )

    # Real table -> banded styling + auto-filter in Excel; safe in Sheets too
    ref = f"A1:{get_column_letter(ncol)}{nrow + 1}"
    table = Table(displayName="WorldCupPlayers", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
    ws.add_table(table)

    # Colour-scale on Total goals (column E)
    goals_col = get_column_letter(headers.index("Total goals") + 1)
    ws.conditional_formatting.add(
        f"{goals_col}2:{goals_col}{nrow + 1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       mid_type="num", mid_value=4, mid_color="9BD6A8",
                       end_type="max", end_color="0B6B3A"))

    # Column widths from content
    widths = [len(h) for h in headers]
    for record in data:
        for c in range(ncol):
            if c < len(record):
                widths[c] = max(widths[c], len(record[c]))
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = min(w + 3, MAX_WIDTH)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26

    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}  ({nrow} players)")


if __name__ == "__main__":
    main()
